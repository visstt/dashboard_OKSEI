#!/usr/bin/env python3
"""
Конвертер XLS → XLSX
Использует библиотеки xlrd и openpyxl для конвертации старых Excel файлов в новый формат
"""

import sys
import os

def convert_xls_to_xlsx(input_file, output_file):
    """Конвертирует XLS файл в XLSX"""
    try:
        # Пробуем использовать xlrd для чтения XLS
        import xlrd
        from openpyxl import Workbook
        
        # Открываем старый XLS файл
        wb_xls = xlrd.open_workbook(input_file)
        
        # Создаём новый XLSX файл
        wb_xlsx = Workbook()
        wb_xlsx.remove(wb_xlsx.active)  # Удаляем дефолтный лист
        
        # Копируем каждый лист
        for sheet_name in wb_xls.sheet_names():
            sheet_xls = wb_xls.sheet_by_name(sheet_name)
            sheet_xlsx = wb_xlsx.create_sheet(title=sheet_name)
            
            # Копируем данные
            for row_idx in range(sheet_xls.nrows):
                for col_idx in range(sheet_xls.ncols):
                    cell = sheet_xls.cell(row_idx, col_idx)
                    cell_xlsx = sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1)
                    
                    # Определяем тип ячейки и копируем значение
                    if cell.ctype == xlrd.XL_CELL_NUMBER:
                        cell_xlsx.value = cell.value
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        # Конвертируем дату Excel в Python datetime
                        from datetime import datetime
                        from xlrd import xldate_as_datetime
                        try:
                            dt = xldate_as_datetime(cell.value, wb_xls.datemode)
                            cell_xlsx.value = dt
                        except:
                            cell_xlsx.value = cell.value
                    else:
                        cell_xlsx.value = cell.value
        
        # Сохраняем XLSX файл
        wb_xlsx.save(output_file)
        print(f" Конвертировано: {input_file} → {output_file}")
        return True
        
    except ImportError as e:
        print(f" Ошибка: не установлены необходимые библиотеки")
        print(f"Установите: pip3 install xlrd openpyxl")
        print(f"Ошибка: {e}")
        return False
    except Exception as e:
        print(f" Ошибка конвертации: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python3 xls_to_xlsx.py <input.xls> [output.xlsx]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f" Файл не найден: {input_file}")
        sys.exit(1)
    
    # Если выходной файл не указан, создаём с тем же именем но расширением .xlsx
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        output_file = os.path.splitext(input_file)[0] + ".xlsx"
    
    success = convert_xls_to_xlsx(input_file, output_file)
    sys.exit(0 if success else 1)
